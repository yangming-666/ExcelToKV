import re
import openpyxl

# ============================================================
# 注释行判断（独立定义，避免跨模块依赖）
# ============================================================

def is_commented_row(row):
    if not row:
        return True
    for cell in row:
        if cell not in (None, "", " "):
            return str(cell).strip().startswith("#")
    return True

def insert_header_by_kv_order(headers, new_field, kv_field_order, pk_col_name):
    """
    headers: 当前 Excel header（list，可包含 None）
    new_field: KV 中新出现的字段
    kv_field_order: KV 中字段的原始顺序
    pk_col_name: PK 列名（headers[pk_col-1]）
    """

    # 找 new_field 在 KV 中的位置
    idx = kv_field_order.index(new_field)

    # 向前找一个“已存在于 Excel header 中的字段”
    insert_after = None
    for i in range(idx - 1, -1, -1):
        prev = kv_field_order[i]
        if prev in headers:
            insert_after = prev
            break

    if insert_after:
        insert_pos = headers.index(insert_after) + 1
    else:
        # 没找到前驱字段：插到 PK 后
        insert_pos = headers.index(pk_col_name) + 1

    headers.insert(insert_pos, new_field)
    return insert_pos



# ============================================================
# KV 解析（顺序幂等核心）
# ============================================================

_QUOTED_RE = re.compile(r'"([^"]*)"')

def _norm_pk(x):
    s = str(x).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s

def parse_kv_simple(path):
    """
    返回：
    root: str
    data: {
        pk: {
            "__field_order__": [field1, field2, ...],
            field: value | dict,
        }
    }
    pk_order: [pk1, pk2, ...]
    global_field_order: [fieldA, fieldB, ...]  # 跨 PK 的全局字段出现顺序（按 KV 原文）
    """

    with open(path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    root = None
    pk_order = []
    data = {}

    # ✅ 新增：全局字段出现顺序（按 KV 原文扫描顺序）
    global_field_order = []
    global_field_set = set()

    stack = []
    pending_key = None

    current_pk = None
    current_block = None

    for raw in lines:
        s = raw.strip()
        if not s:
            continue
        if s.startswith("#") or s.startswith("//"):
            continue

        if s == "{":
            if pending_key is None:
                continue

            key = pending_key
            pending_key = None

            if root is None:
                root = key
                stack.append("root")

            elif current_pk is None:
                current_pk = key
                pk_order.append(current_pk)
                data[current_pk] = {
                    "__field_order__": []
                }
                stack.append("pk")

            else:
                current_block = key
                if current_block not in data[current_pk]:
                    data[current_pk][current_block] = {
                        "__sub_order__": []
                    }
                    data[current_pk]["__field_order__"].append(current_block)

                    # ✅ 记录到全局顺序
                    if current_block not in global_field_set:
                        global_field_set.add(current_block)
                        global_field_order.append(current_block)

                stack.append("block")

            continue

        if s == "}":
            if stack:
                t = stack.pop()
                if t == "block":
                    current_block = None
                elif t == "pk":
                    current_pk = None
            continue

        qs = _QUOTED_RE.findall(s)

        # "Key"
        if len(qs) == 1:
            pending_key = qs[0]
            continue

        # "k" "v"
        if len(qs) >= 2 and current_pk:
            k, v = qs[0], qs[1]

            if current_block:
                block = data[current_pk][current_block]
                if k not in block:
                    block["__sub_order__"].append(k)
                block[k] = v
            else:
                if k not in data[current_pk]:
                    data[current_pk]["__field_order__"].append(k)

                    # ✅ 记录到全局顺序（跨 PK 新字段会在这里首次出现）
                    if k not in global_field_set:
                        global_field_set.add(k)
                        global_field_order.append(k)

                data[current_pk][k] = v

    if root is None:
        raise ValueError("KV 解析失败：找不到 root")

    return root, data, pk_order, global_field_order

# ============================================================
# Excel 结构分析（与 excel_to_kv 完全对齐）
# ============================================================

def analyze_excel_layout(ws):
    rows = list(ws.iter_rows(values_only=True))

    # root
    root_pos = None
    for r, row in enumerate(rows, start=1):
        if is_commented_row(row):
            continue
        for c, v in enumerate(row, start=1):
            if v not in (None, "", " "):
                root_pos = (r, c)
                break
        if root_pos:
            break

    if not root_pos:
        raise ValueError("Excel 中找不到 Root")

    # header
    header_row = None
    for r, row in enumerate(rows, start=1):
        if is_commented_row(row):
            continue
        header_row = r
        break

    headers = [
        str(v).strip() if v not in (None, "", " ") else None
        for v in rows[header_row - 1]
    ]

    pk_col = None
    for i, h in enumerate(headers):
        if h is not None:
            pk_col = i + 1
            break

    if pk_col is None:
        raise ValueError("Excel Header 中找不到 PK 列")

    return {
        "root_pos": root_pos,
        "header_row": header_row,
        "headers": headers,
        "pk_col": pk_col,
        "data_start": header_row + 1,
    }


# ============================================================
# 编码嵌套 block（顺序稳定）
# ============================================================

def encode_nested_block(block: dict):
    order = block.get("__sub_order__", [])
    parts = []
    for k in order:
        parts.append(f"{k}|{block[k]}")
    return ",".join(parts)


# ============================================================
# KV → Excel（顺序幂等最终实现）
# ============================================================

def kv_to_excel_idempotent_sync(kv_path, excel_path):
    kv_root, kv_data, kv_pk_order, global_field_order = parse_kv_simple(kv_path)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.worksheets[0]

    meta = analyze_excel_layout(ws)

    # 1. 覆盖 Root
    r, c = meta["root_pos"]
    ws.cell(row=r, column=c).value = kv_root

    # 2. 当前 Excel header（保持顺序）
    headers = meta["headers"]
    header_set = set(h for h in headers if h)

    pk_col_name = headers[meta["pk_col"] - 1]

    for field in global_field_order:
        if field in header_set:
            continue

        # 仍然按 KV 顺序“插入一列”，不是 append：找前驱字段
        insert_after = None
        idx = global_field_order.index(field)
        for i in range(idx - 1, -1, -1):
            prev = global_field_order[i]
            if prev in headers:
                insert_after = prev
                break

        if insert_after:
            insert_pos = headers.index(insert_after) + 1
        else:
            insert_pos = headers.index(pk_col_name) + 1

        # 内存 header 插入
        headers.insert(insert_pos, field)

        # Excel 物理插列（保持注释/列绑定不漂移）
        ws.insert_cols(insert_pos + 1)  # openpyxl 1-based
        ws.cell(row=meta["header_row"], column=insert_pos + 1).value = field

        header_set.add(field)

    # 5. Excel PK → 行号
    excel_pk_row = {}
    for r in range(meta["data_start"], ws.max_row + 1):
        row_vals = [cell.value for cell in ws[r]]
        if is_commented_row(row_vals):
            continue
        pk_val = ws.cell(row=r, column=meta["pk_col"]).value
        if not pk_val:
            continue
        pk = _norm_pk(pk_val)
        if pk.startswith("#"):
            continue
        excel_pk_row[pk] = r

    # 6. 标记 Excel 中已删除 PK
    for pk, row in excel_pk_row.items():
        if pk not in kv_data:
            ws.cell(row=row, column=meta["pk_col"]).value = f"#DELETED {pk}"

    # 7. header → col
    header_to_col = {
        h: i + 1
        for i, h in enumerate(headers)
        if h and (i + 1) != meta["pk_col"]
    }

    # 8. 写数据（不改变顺序）
    for pk in kv_pk_order:
        fields = kv_data[pk]

        if pk in excel_pk_row:
            row = excel_pk_row[pk]
        else:
            row = ws.max_row + 1
            ws.cell(row=row, column=meta["pk_col"]).value = pk

        for field in headers:
            if field is None or field == headers[meta["pk_col"] - 1]:
                continue

            col = header_to_col.get(field)
            if col is None:
                continue

            val = fields.get(field)
            if isinstance(val, dict):
                ws.cell(row=row, column=col).value = encode_nested_block(val)
            elif val is None:
                ws.cell(row=row, column=col).value = ""
            else:
                ws.cell(row=row, column=col).value = str(val)

    wb.save(excel_path)
